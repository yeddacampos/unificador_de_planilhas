# -*- coding: utf-8 -*-
"""Planilha de Mapeamento.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1k0KEWc2y8eHnQuxgzv0_PgFqj5q98VGx
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google.colab import files

# Caminhos e páginas
arquivo1 = "Caminho do Arquivo 1"
arquivo2 = "Caminho do Arquivo 2"
pagina = "Nome da Página a ser Unificada"

# Carregar a página específica de cada planilha
df1 = pd.read_excel(arquivo1, sheet_name=pagina)
df2 = pd.read_excel(arquivo2, sheet_name=pagina)

# Concatenar os dados das duas páginas
dados_concatenados = pd.concat([df1, df2], ignore_index=True)

# Substituir valores nulos ou vazios por string vazia
dados_concatenados.fillna('', inplace=True)

# Salvar em uma nova planilha
arquivo_destino = "caminho do local a ser salvo"
dados_concatenados.to_excel(arquivo_destino, sheet_name="Nome da Nova Página", index=False)


# Abre o arquivo para aplicar estilos
wb = load_workbook(arquivo_destino)
ws = wb.active  # Carrega a planilha ativa

# Definir estilos para o cabeçalho
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
header_fill = PatternFill(start_color="5071c2", end_color="5071c2", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# Aplicar estilos ao cabeçalho
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = border

# Definir estilos para as linhas
row_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

# Aplicar estilos ao corpo (linhas)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)  # Alinhamento e quebra de texto
        cell.border = border
        cell.fill = row_fill

# Ajustar a largura das colunas
largura_padrao = 25  # Ajuste para o valor que você preferir
for column in ws.columns:
    ws.column_dimensions[column[0].column_letter].width = largura_padrao

# Salvar o arquivo estilizado
wb.save(arquivo_destino)

dados_concatenados.head()

output_path = arquivo_destino

#download
files.download(output_path)